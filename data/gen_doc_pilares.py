# -*- coding: utf-8 -*-
"""Genera el libro Excel de referencia: qué preguntas alimentan cada pilar (usuarios)."""
import json, io, sys, collections

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding="utf-8")
ROOT = "c:/Users/cmonz/Desktop/MLC POC"
OUT = sys.argv[1]

D = json.load(io.open(ROOT + "/data/mock-usuarios.json", encoding="utf-8"))
IND = {i["field"]: i for i in json.load(
    io.open(ROOT + "/data/indicators-usuarios.json", encoding="utf-8"))}
sc, rs = D["schema"], D["responses"]

POS = {"si", "siempre", "optimo", "adecuado", "indetectable", "menos_de_30_minutos",
       "muy_bueno", "bueno", "90porcentaje", "mensual", "si_para_todos",
       "si_periodicamente", "si_de_forma_sistematizada", "si_sistematicamente"}
PAR = {"a_veces", "ocasionalmente", "parcialmente", "regular", "entre_30_y_60_minutos",
       "si_pero_sin_sistematizacion", "30_dias", "60_dias", "90_dias", "trimestral", "semestral"}
NS = {"no_se", "no_recuerda", "no_se_no_recuerda", "no_sabe",
      "prefiero_no_responder", "no_aplica", "", None}
INV = {"p18", "p23", "p25", "p26", "p27", "p36", "p45"}
EXC = {"p34"}
MINQ = 3
ETIQ = {"si": "Sí", "no": "No", "no_se": "No sé", "no_recuerda": "No recuerda",
        "no_se_no_recuerda": "No sé / No recuerda", "prefiero_no_responder": "Prefiero no responder",
        "parcialmente": "Parcialmente", "regular": "Regular", "muy_bueno": "Muy bueno",
        "bueno": "Bueno", "malo": "Malo", "muy_malo": "Muy malo",
        "menos_de_30_minutos": "Menos de 30 minutos",
        "entre_30_y_60_minutos": "Entre 30 y 60 minutos",
        "mas_de_1_hora": "Más de 1 hora", "detectable": "Detectable",
        "indetectable": "Indetectable"}

PILARES = [
    ("Disponibilidad", "¿Existen los medicamentos, pruebas e insumos de prevención?"),
    ("Accesibilidad", "¿Pueden las personas acceder a los servicios disponibles?"),
    ("Aceptabilidad", "¿Son culturalmente aceptados los servicios?"),
    ("Adecuación", "¿Se ajustan los servicios a las necesidades de las personas?"),
    ("Asequibilidad", "¿Son asequibles y sostenibles los servicios?"),
]
OTROS = [
    ("Cascada", "Continuidad del tratamiento y resultado virológico. NO se muestra en el panel."),
    ("Caracterización", "Perfil de la persona entrevistada. Alimenta filtros y gráficas de perfil."),
    ("Ubicación", "Dónde se levantó la entrevista y desde dónde viaja la persona."),
    ("Cierre", "Respuestas abiertas, se consultan en el modal «Ver respuestas»."),
    ("Consentimiento", "Filtro de participación."),
    ("Fecha", "Fecha de recolección."),
]
DESC = dict(PILARES + OTROS)
ORDEN = [p for p, _ in PILARES] + [p for p, _ in OTROS]

# ---- paleta del proyecto (KEY VISUAL) ----
OCEANIC, CREAM, CRIMSON = "FF003049", "FFFDF0D5", "FFC1111F"
GREY_BG, ZEBRA = "FFEBF1F3", "FFF7FAFB"
V_FULL, V_HALF, V_ZERO, V_NS = "FFDDEFEC", "FFF6E9D2", "FFF6DCDE", "FFE5ECEE"
T_FULL, T_HALF, T_ZERO, T_NS = "FF14746A", "FF8F6210", "FFB00E1B", "FF6D818C"

thin = Side(style="thin", color="FFD5DFE4")
BORDE = Border(bottom=thin)


def contestada(v):
    return not (isinstance(v, list) or v in NS)


def n_de(name):
    return sum(1 for r in rs if contestada(r.get(name)))


def rol(q):
    """(¿puntúa?, rol, sentido, motivo)"""
    n = n_de(q["name"])
    if q["name"] in EXC:
        return False, "Excluida", "", "No mide calidad del servicio (pregunta de tamizaje)"
    if not str(q["type"]).startswith("select_one"):
        t = str(q["type"]).split()[0]
        nombre = {"text": "Texto abierto", "decimal": "Numérica", "integer": "Numérica",
                  "date": "Fecha", "select_multiple": "Opción múltiple"}.get(t, t)
        return False, "No puntúa", "", "No es de opción única (%s)" % nombre
    if n < MINQ:
        return False, "Excluida", "", "La contestaron menos de %d personas (n=%d)" % (MINQ, n)
    if q["name"] in INV:
        return True, "Puntúa", "Invertido", ""
    return True, "Puntúa", "Directo", ""


def valor(v, name):
    if v in NS or isinstance(v, list):
        return None
    base = 1.0 if v in POS else 0.5 if v in PAR else 0.0
    return (1.0 - base) if name in INV else base


def tipo_legible(q):
    t = str(q["type"]).split()[0]
    return {"text": "Texto abierto", "decimal": "Numérica", "integer": "Numérica",
            "date": "Fecha", "select_multiple": "Opción múltiple",
            "select_one": "Opción única"}.get(t, t)


wb = openpyxl.Workbook()


def encabezado(ws, titulos, anchos, fila=1):
    for i, (t, w) in enumerate(zip(titulos, anchos), start=1):
        c = ws.cell(row=fila, column=i, value=t)
        c.font = Font(bold=True, color=CREAM, size=10)
        c.fill = PatternFill("solid", fgColor=OCEANIC)
        c.alignment = Alignment(vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[fila].height = 30
    ws.freeze_panes = ws.cell(row=fila + 1, column=1)
    ws.auto_filter.ref = "A%d:%s%d" % (fila, get_column_letter(len(titulos)), fila)


# ============================================================ 1. Resumen
ws = wb.active
ws.title = "Resumen por pilar"
ws["A1"] = "Pilares y preguntas · Encuesta de usuarios MLC Guatemala"
ws["A1"].font = Font(bold=True, size=15, color=OCEANIC)
ws["A2"] = ("Qué pregunta alimenta cada pilar del panel, cuáles entran al puntaje y con qué valor "
            "cuenta cada respuesta. Elaborado sobre las %d entrevistas cargadas en el tablero." % len(rs))
ws["A2"].font = Font(size=10, color="FF33505F")
ws["A2"].alignment = Alignment(wrap_text=True)
ws.merge_cells("A1:F1")
ws.merge_cells("A2:F2")
ws.row_dimensions[2].height = 28

encabezado(ws, ["Pilar o grupo", "Qué mide", "¿Se muestra en el panel?", "Preguntas",
                "Puntúan", "De sentido invertido", "Excluidas o sin puntaje"],
           [20, 52, 20, 11, 10, 18, 20], fila=4)

# Los grupos que no son pilares (Caracterización, Ubicación, Cierre...) nunca se
# puntúan, así que sus conteos irían vacíos en vez de sugerir un cálculo que no ocurre.
PILAR_NOMBRES = dict(PILARES)
fila = 5
for pilar in ORDEN:
    qs = [q for q in sc if q["pilar"] == pilar]
    if not qs:
        continue
    es_pilar = pilar in PILAR_NOMBRES
    if es_pilar:
        estado = "Sí, pestaña propia"
    elif pilar == "Cascada":
        estado = "No, falta la pestaña"
    else:
        estado = "No es un pilar"
    calcula = es_pilar or pilar == "Cascada"
    p = sum(1 for q in qs if rol(q)[0])
    inv = sum(1 for q in qs if rol(q)[0] and q["name"] in INV)
    vals = [pilar, DESC.get(pilar, ""), estado, len(qs),
            p if calcula else "—", inv if calcula else "—",
            (len(qs) - p) if calcula else "—"]
    for i, v in enumerate(vals, start=1):
        c = ws.cell(row=fila, column=i, value=v)
        c.border = BORDE
        c.alignment = Alignment(vertical="top", wrap_text=(i == 2),
                                horizontal="center" if i >= 4 else "left")
        c.font = Font(size=10, color="FF33505F")
        if i == 1:
            c.font = Font(bold=es_pilar, color=OCEANIC if es_pilar else "FF33505F", size=10)
        if i == 3:
            if es_pilar:
                c.fill = PatternFill("solid", fgColor=V_FULL)
                c.font = Font(size=10, bold=True, color=T_FULL)
            elif pilar == "Cascada":
                c.fill = PatternFill("solid", fgColor=V_ZERO)
                c.font = Font(size=10, bold=True, color=T_ZERO)
            else:
                c.fill = PatternFill("solid", fgColor=V_NS)
                c.font = Font(size=10, color=T_NS)
        elif not es_pilar:
            c.fill = PatternFill("solid", fgColor=ZEBRA)
    ws.row_dimensions[fila].height = 28
    fila += 1

fila += 1
for txt in ["Los cinco pilares del panel van primero. «Cascada» sí se calcula pero no tiene pestaña, "
            "así que hoy no se ve.",
            "Los grupos restantes no son pilares: alimentan filtros, gráficas de perfil y respuestas "
            "abiertas, y por eso sus columnas de puntaje van en blanco."]:
    c = ws.cell(row=fila, column=1, value=txt)
    c.font = Font(italic=True, size=9, color="FF6B818C")
    ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=7)
    fila += 1

# ============================================================ 2. Preguntas
ws2 = wb.create_sheet("Preguntas")
encabezado(ws2, ["Pilar", "Campo", "Pregunta", "Indicador", "Tipo de respuesta", "n",
                 "Rol en el puntaje", "Sentido", "Por qué no puntúa"],
           [17, 8, 62, 34, 16, 6, 15, 11, 42])

fila = 2
for pilar in ORDEN:
    for q in [q for q in sc if q["pilar"] == pilar]:
        ok, r, sentido, motivo = rol(q)
        vals = [pilar, q["name"], q["label"], IND.get(q["name"], {}).get("name", ""),
                tipo_legible(q), n_de(q["name"]), r, sentido, motivo]
        for i, v in enumerate(vals, start=1):
            c = ws2.cell(row=fila, column=i, value=v)
            c.border = BORDE
            c.font = Font(size=10, color="FF0D1F2A" if ok else "FF6B818C")
            c.alignment = Alignment(vertical="top", wrap_text=(i in (3, 4, 9)))
            if i == 2:
                c.font = Font(size=10, bold=True, color=OCEANIC)
            if i == 7:
                if r == "Puntúa":
                    c.fill = PatternFill("solid", fgColor=V_FULL)
                    c.font = Font(size=10, bold=True, color=T_FULL)
                elif r == "Excluida":
                    c.fill = PatternFill("solid", fgColor=V_ZERO)
                    c.font = Font(size=10, bold=True, color=T_ZERO)
                else:
                    c.fill = PatternFill("solid", fgColor=V_NS)
                    c.font = Font(size=10, color=T_NS)
            if i == 8 and sentido == "Invertido":
                c.fill = PatternFill("solid", fgColor=V_HALF)
                c.font = Font(size=10, bold=True, color=T_HALF)
        ws2.row_dimensions[fila].height = 30
        fila += 1

# ============================================================ 3. Opciones
ws3 = wb.create_sheet("Opciones de respuesta")
encabezado(ws3, ["Pilar", "Campo", "Pregunta", "Sentido", "Opción de respuesta",
                 "Respuestas", "% del total", "Valor", "Cuenta como"],
           [17, 8, 52, 11, 26, 11, 11, 8, 16])

fila = 2
for pilar in ORDEN:
    for q in [q for q in sc if q["pilar"] == pilar]:
        if not str(q["type"]).startswith("select_one"):
            continue
        ok, r, sentido, motivo = rol(q)
        c_opt = collections.Counter(rr.get(q["name"]) for rr in rs)
        total = sum(c_opt.values())
        for v, n in c_opt.most_common():
            val = valor(v, q["name"])
            etiqueta = ETIQ.get(v, (v or "(sin respuesta)").replace("_", " ").capitalize())
            if val is None:
                cuenta, fill, color = "Fuera del denominador", V_NS, T_NS
            elif val == 1.0:
                cuenta, fill, color = "Favorable", V_FULL, T_FULL
            elif val == 0.5:
                cuenta, fill, color = "Parcial", V_HALF, T_HALF
            else:
                cuenta, fill, color = "Desfavorable", V_ZERO, T_ZERO
            vals = [pilar, q["name"], q["label"], sentido or "—", etiqueta, n,
                    n / total if total else 0, val if val is not None else "NS", cuenta]
            for i, x in enumerate(vals, start=1):
                c = ws3.cell(row=fila, column=i, value=x)
                c.border = BORDE
                c.font = Font(size=10, color="FF33505F")
                c.alignment = Alignment(vertical="top", wrap_text=(i == 3))
                if i == 2:
                    c.font = Font(size=10, bold=True, color=OCEANIC)
                if i == 7:
                    c.number_format = "0.0%"
                if i in (8, 9):
                    c.fill = PatternFill("solid", fgColor=fill)
                    c.font = Font(size=10, bold=(i == 8), color=color)
                    c.alignment = Alignment(horizontal="center" if i == 8 else "left",
                                            vertical="top")
            fila += 1

# ============================================================ 4. Cómo se calcula
ws4 = wb.create_sheet("Cómo se calcula")
ws4.column_dimensions["A"].width = 4
ws4.column_dimensions["B"].width = 104

filas = [
    ("t", "Cómo se calcula el puntaje de un pilar"),
    ("", ""),
    ("h", "Las tres reglas"),
    ("n", "1. El conjunto de preguntas se elige una sola vez, sobre las %d entrevistas completas y "
          "no por establecimiento. Así los puntajes siguen siendo comparables aunque se filtre o se "
          "cambie de pestaña en el panel." % len(rs)),
    ("n", "2. Cada pregunta se promedia entre quienes la contestaron y después se promedian las "
          "preguntas entre sí. Todas pesan igual: sin esto, una pregunta que el formulario le mostró "
          "a más gente dominaría el pilar sin que eso fuera una decisión metodológica."),
    ("n", "3. El pilar es el promedio simple de sus preguntas. Una pregunta que nadie del "
          "establecimiento contestó no entra, y si ninguna tiene respuestas el pilar se muestra "
          "como «—» en vez de cero."),
    ("", ""),
    ("h", "Valor de cada respuesta"),
    ("n", "Favorable = 1 · Parcial = 0.5 · Desfavorable = 0."),
    ("n", "«No sé», «Prefiero no responder» y las preguntas que el formulario saltó quedan fuera del "
          "denominador: no suman ni restan."),
    ("", ""),
    ("h", "Preguntas de sentido invertido"),
    ("n", "En estas el «Sí» es la respuesta desfavorable, así que el valor se invierte (1 − valor) "
          "antes de promediar:"),
]
for name in sorted(INV):
    q = next((x for x in sc if x["name"] == name), None)
    if q:
        filas.append(("l", "%s — %s" % (name, q["label"])))
filas += [
    ("", ""),
    ("h", "Preguntas excluidas del puntaje"),
    ("n", "p34 «¿Ha consultado por alguna infección de transmisión sexual?» es tamizaje: que alguien "
          "haya consultado no dice nada bueno ni malo del servicio."),
    ("n", "p36 quedó fuera porque la contestó una sola persona, por debajo del mínimo de %d." % MINQ),
    ("", ""),
    ("h", "Advertencia sobre el pilar «Cascada»"),
    ("n", "Sus cuatro preguntas se recolectan y tienen indicador definido, pero el panel sólo tiene "
          "cinco pestañas de pilar y «Cascada» no es una de ellas, así que sus tarjetas nunca se "
          "dibujan. Para verlas habría que agregar la pestaña."),
    ("", ""),
    ("h", "Fuente"),
    ("n", "data/mock-usuarios.json y data/indicators-usuarios.json. Las listas de preguntas "
          "invertidas y excluidas viven en js/dashboard-data.js (MLC.sentidoInverso y MLC.sinPuntaje)."),
]

fila = 1
for tipo, txt in filas:
    c = ws4.cell(row=fila, column=2, value=txt)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    if tipo == "t":
        c.font = Font(bold=True, size=15, color=OCEANIC)
    elif tipo == "h":
        c.font = Font(bold=True, size=11, color=CRIMSON)
    elif tipo == "l":
        c.font = Font(size=10, color="FF33505F")
        ws4.cell(row=fila, column=1, value="·").font = Font(size=10, color=T_HALF)
    else:
        c.font = Font(size=10, color="FF33505F")
    ws4.row_dimensions[fila].height = 15 if not txt else (34 if len(txt) > 110 else 17)
    fila += 1

for hoja in wb.worksheets:
    hoja.sheet_view.showGridLines = False

wb.save(OUT)
print("generado:", OUT)
print("hojas:", ", ".join(w.title for w in wb.worksheets))
print("preguntas:", sum(1 for q in sc), "| filas de opciones:", ws3.max_row - 1)
