"""
Generador de datos mock para los dashboards MLC.

Lee la estructura de las encuestas en `info/kobo_prestadores.xlsx` y
`info/kobo_usuarios.xlsx`, genera 50 respuestas simuladas por encuesta
y guarda los resultados como JSON consumible por los dashboards.

El campo `pilar` ya no vive en las hojas de Kobo (solo `type`/`name`/`label`),
así que se cruza con `info/indicadores.xlsx` (hojas "Usuario" y
"Prestadores de servicios ") usando el ID de pregunta normalizado.

Salidas:
    data/mock-prestadores.json
    data/mock-usuarios.json
"""

from __future__ import annotations
import json
import random
import re
from datetime import date, timedelta
from pathlib import Path

import openpyxl

random.seed(42)

ROOT = Path(__file__).resolve().parent.parent
INFO = ROOT / "info"
OUT = ROOT / "data"

ETNIAS = ["maya", "ladino_o_mestizo", "xinka", "garifuna", "otro"]


# Pool fijo de unidades para el dashboard de Prestadores (mantenidas tal cual
# se mostraban en la maqueta original — coinciden con el filtro hardcodeado).
PRESTADOR_UNIDADES = [
    {"nombre": "Centro de Salud Roosevelt",          "departamento": "GUATEMALA",      "municipio": "GUATEMALA",       "tipo": "Centro de salud"},
    {"nombre": "Centro de Salud San Juan de Dios",   "departamento": "GUATEMALA",      "municipio": "GUATEMALA",       "tipo": "Centro de salud"},
    {"nombre": "Puesto de Salud Esquintla",          "departamento": "ESCUINTLA",      "municipio": "ESCUINTLA",       "tipo": "Puesto de salud"},
    {"nombre": "Puesto de Salud Quetzaltenango",     "departamento": "QUETZALTENANGO", "municipio": "QUETZALTENANGO",  "tipo": "Puesto de salud"},
    {"nombre": "Puesto de Salud Izabal",             "departamento": "IZABAL",         "municipio": "PUERTO BARRIOS",  "tipo": "Puesto de salud"},
]


# ---------------------------------------------------------------------------
# Catálogo de unidades (hoja UNIDADES de indicadores.xlsx)
# ---------------------------------------------------------------------------
def load_unidades() -> list[dict]:
    """Devuelve [{departamento, municipio, nombre, tipo, codigo}] desde la hoja UNIDADES."""
    wb = openpyxl.load_workbook(INFO / "indicadores.xlsx", data_only=True)
    ws = wb["UNIDADES"]
    rows = []
    for r in ws.iter_rows(values_only=True):
        if all(v is None for v in r):
            break
        rows.append(r)
    if not rows:
        return []
    out = []
    for r in rows[1:]:
        nombre = r[2]
        if not nombre:
            continue
        out.append({
            "departamento": str(r[0] or "").strip(),
            "municipio":    str(r[1] or "").strip(),
            "nombre":       str(nombre).strip(),
            "tipo":         str(r[3] or "").strip(),
            "codigo":       str(r[4] or "").strip(),
        })
    # Normaliza inconsistencias mayúsculas/minúsculas en departamento
    for u in out:
        u["departamento"] = u["departamento"].upper()
    return out


def pick_sample_unidades(unidades: list[dict], n: int = 12) -> list[dict]:
    """Selecciona n unidades de forma diversa: distintos departamentos y tipos."""
    rng = random.Random(42)
    by_dept: dict[str, list[dict]] = {}
    for u in unidades:
        if not u["departamento"] or not u["municipio"]:
            continue
        by_dept.setdefault(u["departamento"], []).append(u)
    depts = list(by_dept.keys())
    rng.shuffle(depts)
    sample = []
    while len(sample) < n and any(by_dept.values()):
        for d in depts:
            if not by_dept[d]:
                continue
            u = rng.choice(by_dept[d])
            sample.append(u)
            by_dept[d] = [x for x in by_dept[d] if x["nombre"] != u["nombre"]]
            if len(sample) >= n:
                break
    return sample[:n]


# ---------------------------------------------------------------------------
# Cruce con indicadores.xlsx para obtener el pilar por pregunta
# ---------------------------------------------------------------------------
def _name_to_id(name: str) -> str:
    """'p02_1' -> 'P02.1', 'p18' -> 'P18', 'P03' -> 'P03'."""
    if not name:
        return ""
    s = str(name).strip()
    m = re.match(r"^([Pp]\d+)(?:_(\d+))?(?:_.*)?$", s)
    if m and m.group(2):
        return f"{m.group(1).upper()}.{m.group(2)}"
    m2 = re.match(r"^([Pp]\d+)(_|$)", s)
    if m2:
        return m2.group(1).upper()
    return s.upper()


def load_pilar_map(sheet_name: str, id_col_label: str) -> dict[str, str]:
    """Devuelve {ID_normalizado: pilar} desde indicadores.xlsx."""
    wb = openpyxl.load_workbook(INFO / "indicadores.xlsx", data_only=True)
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    header = [str(h).strip() if h else "" for h in rows[0]]

    def find(label, fuzzy=False):
        l = label.lower()
        for i, h in enumerate(header):
            hl = h.lower()
            if (hl == l) or (fuzzy and l in hl):
                return i
        return None

    i_pilar = find("Pilar")
    i_id    = find(id_col_label) or find(id_col_label, fuzzy=True)
    if i_pilar is None or i_id is None:
        return {}

    out: dict[str, str] = {}
    for r in rows[1:]:
        pilar = r[i_pilar]
        pcode = r[i_id]
        if not pcode or not pilar:
            continue
        out[str(pcode).strip().upper()] = str(pilar).strip()
    return out


# ---------------------------------------------------------------------------
# Carga de la estructura de la encuesta desde un xlsx KoboToolbox
# ---------------------------------------------------------------------------
def load_kobo(path: Path):
    wb = openpyxl.load_workbook(path, data_only=True)
    sheet = wb["survey"]
    header = [c.value for c in sheet[1]]
    i_type  = header.index("type")
    i_name  = header.index("name")
    i_label = next(i for i, h in enumerate(header) if h and "label" in str(h).lower())

    questions = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row or not row[i_type]:
            continue
        t = str(row[i_type]).strip()
        if t.startswith(("begin", "end")) or t in ("note", "calculate"):
            continue
        questions.append({
            "type":  t,
            "name":  row[i_name],
            "label": row[i_label],
        })

    sheet = wb["choices"]
    header = [c.value for c in sheet[1]]
    i_list  = header.index("list_name")
    i_name  = header.index("name")
    i_label = next(i for i, h in enumerate(header) if h and "label" in str(h).lower())
    choices: dict[str, list[dict]] = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row or not row[i_list]:
            continue
        ln = str(row[i_list]).strip()
        choices.setdefault(ln, []).append({"name": row[i_name], "label": row[i_label]})

    return questions, choices


def attach_pilar(questions: list[dict], pilar_map: dict[str, str]) -> list[dict]:
    """Devuelve copia de `questions` con un campo `pilar` poblado desde indicadores.xlsx."""
    out = []
    for q in questions:
        pcode = _name_to_id(q.get("name") or "")
        pilar = pilar_map.get(pcode) or None
        out.append({**q, "pilar": pilar})
    return out


# ---------------------------------------------------------------------------
# Generación de respuestas
# ---------------------------------------------------------------------------
def list_name(q_type: str) -> str | None:
    parts = q_type.split()
    if parts[0] in ("select_one", "select_multiple") and len(parts) >= 2:
        return parts[1]
    return None


def weighted_choice(options: list[str], biases: dict[str, float] | None = None) -> str:
    if biases is None:
        biases = {}
    weights = []
    for opt in options:
        if opt in biases:
            weights.append(biases[opt])
        elif opt in ("si", "siempre", "si_para_todos", "si_periodicamente",
                     "si_de_forma_sistematizada", "si_sistematicamente"):
            weights.append(0.55)
        elif opt in ("a_veces", "ocasionalmente", "parcialmente",
                     "si_pero_sin_sistematizacion", "solo_ocasional"):
            weights.append(0.18)
        elif opt in ("no",):
            weights.append(0.18)
        elif opt in ("no_se", "no_se_no_recuerda", "prefiero_no_responder"):
            weights.append(0.05)
        else:
            weights.append(0.1)
    return random.choices(options, weights=weights, k=1)[0]


def random_date(days_back: int = 365) -> str:
    delta = timedelta(days=random.randint(0, days_back))
    return (date.today() - delta).isoformat()


SI_NO_NS = ["si", "no", "no_se"]


def looks_like_yes_no(label: str) -> bool:
    if not label:
        return False
    s = label.strip().lower()
    return s.startswith(("¿", "se ", "el ", "la ", "los ", "las ", "existe", "cuenta", "ofertan",
                         "cuenta con", "se realiza", "se asigna", "se entrega"))


def gen_response(q: dict, choices: dict[str, list[dict]]):
    t = q["type"]

    if t == "date":
        return random_date()

    if t == "integer":
        label = (q.get("label") or "").lower()
        if "minuto" in label or ("tiempo" in label and "promedio" in label):
            return random.randint(15, 90)
        if "hora" in label:
            return random.choice([1, 2, 3, 4])
        if "cantidad" in label or "cuánt" in label or "cuant" in label:
            return random.randint(0, 30)
        return random.randint(0, 100)

    if t == "decimal":
        return round(random.uniform(0, 100), 1)

    if t == "text":
        return ""

    if t.startswith("select_one"):
        lname = list_name(t)
        opts = [c["name"] for c in choices.get(lname, [])]
        establishment_like = {"uai", "vicits", "cap", "caimi",
                              "puesto_de_salud", "centro_de_salud", "hospital"}
        if opts and set(opts).issubset(establishment_like) and looks_like_yes_no(q.get("label") or ""):
            return weighted_choice(SI_NO_NS)
        if not opts:
            return None
        return weighted_choice(opts)

    if t.startswith("select_multiple"):
        lname = list_name(t)
        opts = [c["name"] for c in choices.get(lname, [])]
        if not opts:
            return []
        k = random.randint(1, min(3, len(opts)))
        return random.sample(opts, k)

    return None


# ---------------------------------------------------------------------------
# Generadores específicos por encuesta
# ---------------------------------------------------------------------------
def gen_prestador_response(idx: int, questions, choices, unidades_pool: list[dict]):
    answers = {q["name"]: gen_response(q, choices) for q in questions}
    answers["P01"] = random_date(180)
    u = random.choice(unidades_pool)
    answers["P02"] = u["nombre"]
    return {
        "_id": f"P{idx:03d}",
        "_unidad":        u["nombre"],
        "_departamento":  u["departamento"],
        "_municipio":     u["municipio"],
        "_tipo":          u["tipo"],
        "_fecha": answers["P01"],
        **answers,
    }


def gen_usuario_response(idx: int, questions, choices, unidades_pool: list[dict]):
    answers = {q["name"]: gen_response(q, choices) for q in questions}
    answers["p01"] = random_date(180)
    u = random.choice(unidades_pool)
    # p03 = nombre de establecimiento (texto libre); poblado desde UNIDADES.
    answers["p03"] = u["nombre"]
    answers["p10"] = random.choice(ETNIAS)
    return {
        "_id": f"U{idx:03d}",
        "_unidad":        u["nombre"],
        "_departamento":  u["departamento"],
        "_municipio":     u["municipio"],
        "_tipo":          u["tipo"],
        "_fecha": answers["p01"],
        "_etnia": answers["p10"],
        **answers,
    }


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def schema_record(q: dict) -> dict:
    return {
        "name":  q["name"],
        "label": q["label"],
        "type":  q["type"],
        "pilar": q.get("pilar"),
    }


def main():
    OUT.mkdir(parents=True, exist_ok=True)

    # --- Catálogo maestro de unidades ---
    unidades = load_unidades()
    (OUT / "unidades.json").write_text(
        json.dumps(unidades, ensure_ascii=False, indent=2), encoding="utf-8",
    )
    pool = pick_sample_unidades(unidades, n=12)
    print(f"[ok] data/unidades.json — {len(unidades)} unidades catalogadas (pool mock: {len(pool)})")

    # --- Prestadores ---
    p_q, p_c = load_kobo(INFO / "kobo_prestadores.xlsx")
    p_pilar = load_pilar_map("Prestadores de servicios ", "ID pregunta")
    p_q = attach_pilar(p_q, p_pilar)
    # Prestadores usa un pool fijo de 5 unidades (centros y puestos de salud)
    # para preservar el resumen por unidad y el filtro original.
    prestadores = [gen_prestador_response(i + 1, p_q, p_c, PRESTADOR_UNIDADES) for i in range(50)]
    schema_p = [schema_record(q) for q in p_q]
    (OUT / "mock-prestadores.json").write_text(
        json.dumps({"schema": schema_p, "responses": prestadores}, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    matched_p = sum(1 for q in p_q if q.get("pilar"))
    print(f"[ok] data/mock-prestadores.json — {len(prestadores)} respuestas | {matched_p}/{len(p_q)} preguntas con pilar")

    # --- Usuarios ---
    u_q, u_c = load_kobo(INFO / "kobo_usuarios.xlsx")
    u_pilar = load_pilar_map("Usuario", "ID Pregunta Nuevo")
    u_q = attach_pilar(u_q, u_pilar)
    usuarios = [gen_usuario_response(i + 1, u_q, u_c, pool) for i in range(50)]
    schema_u = [schema_record(q) for q in u_q]
    (OUT / "mock-usuarios.json").write_text(
        json.dumps({"schema": schema_u, "responses": usuarios}, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    matched_u = sum(1 for q in u_q if q.get("pilar"))
    print(f"[ok] data/mock-usuarios.json — {len(usuarios)} respuestas | {matched_u}/{len(u_q)} preguntas con pilar")


if __name__ == "__main__":
    main()
